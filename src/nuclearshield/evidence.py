from __future__ import annotations

import csv
import json
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler

SUPPORTED_EXTENSIONS = {".csv", ".json", ".jsonl", ".xlsx"}

DOMAIN_HINTS: dict[str, tuple[str, ...]] = {
    "OT / SCADA": (
        "scada", "plc", "ot", "protocol", "telemetry", "network", "packet", "connection",
        "pressure", "temperature", "flow", "power", "flux", "configuration", "drift", "sensor",
    ),
    "SAFETY / INTEGRITY": (
        "safety", "integrity", "instrument", "firmware", "baseline", "self_test", "selftest",
        "checksum", "hash", "trip", "protection", "validation", "reliability",
    ),
    "MC&A / SAFEGUARDS": (
        "material", "inventory", "accounting", "mca", "mc_a", "reconciliation", "safeguard",
        "seal", "balance", "variance", "book_inventory", "measured_inventory",
    ),
    "ACCESS / IDENTITY": (
        "access", "badge", "identity", "subject", "user", "door", "entry", "denied", "granted",
        "after_hours", "failed_attempt", "behavior", "insider", "authentication",
    ),
    "COMPLIANCE / AUDIT": (
        "audit", "compliance", "control", "evidence", "inspection", "approval", "authorized",
        "change_request", "policy", "finding", "exception",
    ),
}

TIMESTAMP_HINTS = ("timestamp", "time", "datetime", "date", "event_time", "observed_at", "created_at")


def _normalise_name(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return value


def _as_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        number = float(text)
        return number if math.isfinite(number) else None
    except ValueError:
        return None


@dataclass
class DatasetAnalysis:
    path: Path
    rows: list[dict[str, Any]]
    columns: list[str]
    normalized_columns: list[str]
    domain: str
    domain_confidence: float
    domain_scores: dict[str, float]
    numeric_columns: list[str]
    timestamp_column: str | None
    missing_cells: int
    total_cells: int
    anomaly_score: float = 0.0
    anomaly_count: int = 0
    analyzed_rows: int = 0
    usable_features: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def filename(self) -> str:
        return self.path.name

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def missing_pct(self) -> float:
        return (self.missing_cells / self.total_cells * 100.0) if self.total_cells else 0.0


@dataclass
class EvidencePackage:
    datasets: list[DatasetAnalysis]

    @property
    def total_rows(self) -> int:
        return sum(dataset.row_count for dataset in self.datasets)

    @property
    def overall_risk_score(self) -> float:
        if not self.datasets:
            return 0.0
        weights = [max(1, dataset.analyzed_rows) for dataset in self.datasets]
        return float(np.average([dataset.anomaly_score for dataset in self.datasets], weights=weights))

    @property
    def risk_level(self) -> str:
        score = self.overall_risk_score
        if score >= 70:
            return "HIGH"
        if score >= 40:
            return "ELEVATED"
        return "NORMAL"


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_json(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".jsonl":
        rows: list[dict[str, Any]] = []
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                if line.strip():
                    item = json.loads(line)
                    if isinstance(item, dict):
                        rows.append(item)
        return rows

    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
                return list(value)
        return [payload]
    return []


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires openpyxl. Reinstall NuclearShield dependencies.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration:
        return []
    columns = [str(value).strip() if value is not None else f"column_{index + 1}" for index, value in enumerate(header)]
    rows = []
    for values in iterator:
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({columns[index]: value if index < len(values) else None for index, value in enumerate(values)})
    workbook.close()
    return rows


def read_rows(path: str | Path) -> list[dict[str, Any]]:
    candidate = Path(path).expanduser().resolve()
    if not candidate.exists() or not candidate.is_file():
        raise FileNotFoundError(f"Evidence file not found: {candidate}")
    suffix = candidate.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported evidence type {suffix or '<none>'}: {candidate.name}")
    if suffix == ".csv":
        return _read_csv(candidate)
    if suffix in {".json", ".jsonl"}:
        return _read_json(candidate)
    return _read_xlsx(candidate)


def _all_columns(rows: list[dict[str, Any]]) -> list[str]:
    seen: dict[str, None] = {}
    for row in rows:
        for key in row:
            seen.setdefault(str(key), None)
    return list(seen)


def infer_domain(columns: Iterable[str]) -> tuple[str, float, dict[str, float]]:
    normalized = [_normalise_name(column) for column in columns]
    scores: dict[str, float] = {}
    for domain, hints in DOMAIN_HINTS.items():
        score = 0.0
        for column in normalized:
            tokens = set(column.split("_"))
            for hint in hints:
                if hint == column:
                    score += 3.0
                elif hint in tokens:
                    score += 2.0
                elif len(hint) >= 5 and hint in column:
                    score += 1.0
        scores[domain] = score

    best_domain, best_score = max(scores.items(), key=lambda item: item[1])
    total = sum(scores.values())
    if best_score <= 0:
        return "GENERAL / UNKNOWN", 0.0, scores
    confidence = min(99.0, 45.0 + 55.0 * (best_score / max(total, best_score)))
    return best_domain, confidence, scores


def _find_numeric_columns(rows: list[dict[str, Any]], columns: list[str]) -> list[str]:
    numeric: list[str] = []
    for column in columns:
        values = [_as_float(row.get(column)) for row in rows]
        present = [value for value in values if value is not None]
        if len(present) >= max(10, int(len(rows) * 0.6)) and len(set(round(value, 8) for value in present)) > 1:
            numeric.append(column)
    return numeric


def _find_timestamp_column(columns: list[str]) -> str | None:
    normalized = {_normalise_name(column): column for column in columns}
    for hint in TIMESTAMP_HINTS:
        if hint in normalized:
            return normalized[hint]
    for normalized_name, original in normalized.items():
        if "time" in normalized_name or normalized_name.endswith("date"):
            return original
    return None


def _analyze_numeric(rows: list[dict[str, Any]], numeric_columns: list[str]) -> tuple[float, int, int, list[str], list[str]]:
    if not numeric_columns:
        return 0.0, 0, 0, [], ["No sufficiently complete variable numeric fields were available for anomaly modeling."]

    matrix: list[list[float]] = []
    for row in rows:
        vector = [_as_float(row.get(column)) for column in numeric_columns]
        if all(value is not None for value in vector):
            matrix.append([float(value) for value in vector if value is not None])

    if len(matrix) < 20:
        return 0.0, 0, len(matrix), numeric_columns, ["Too few complete numeric rows for IsolationForest; statistical profiling only."]

    data = np.asarray(matrix, dtype=float)
    scaler = StandardScaler()
    scaled = scaler.fit_transform(data)
    model = IsolationForest(n_estimators=160, contamination="auto", random_state=132)
    predictions = model.fit_predict(scaled)
    raw = -model.score_samples(scaled)
    lo, hi = float(raw.min()), float(raw.max())
    normalized = np.zeros_like(raw) if hi <= lo else (raw - lo) / (hi - lo) * 100.0
    anomaly_count = int(np.sum(predictions == -1))
    # Robust package-level score: upper-tail anomaly pressure rather than a single extreme point.
    risk_score = float(np.percentile(normalized, 90))
    return risk_score, anomaly_count, len(matrix), numeric_columns, []


def analyze_file(path: str | Path) -> DatasetAnalysis:
    candidate = Path(path).expanduser().resolve()
    rows = read_rows(candidate)
    columns = _all_columns(rows)
    normalized_columns = [_normalise_name(column) for column in columns]
    domain, confidence, domain_scores = infer_domain(columns)
    numeric_columns = _find_numeric_columns(rows, columns)
    timestamp_column = _find_timestamp_column(columns)
    total_cells = len(rows) * len(columns)
    missing_cells = sum(
        1 for row in rows for column in columns
        if row.get(column) is None or str(row.get(column)).strip() == ""
    )
    score, anomaly_count, analyzed_rows, usable_features, notes = _analyze_numeric(rows, numeric_columns)
    if not rows:
        notes.append("File contains no tabular records.")
    if domain == "GENERAL / UNKNOWN":
        notes.append("Domain could not be inferred safely from the supplied schema.")
    return DatasetAnalysis(
        path=candidate,
        rows=rows,
        columns=columns,
        normalized_columns=normalized_columns,
        domain=domain,
        domain_confidence=confidence,
        domain_scores=domain_scores,
        numeric_columns=numeric_columns,
        timestamp_column=timestamp_column,
        missing_cells=missing_cells,
        total_cells=total_cells,
        anomaly_score=score,
        anomaly_count=anomaly_count,
        analyzed_rows=analyzed_rows,
        usable_features=usable_features,
        notes=notes,
    )


def analyze_files(paths: Iterable[str | Path]) -> EvidencePackage:
    datasets = [analyze_file(path) for path in paths]
    return EvidencePackage(datasets=datasets)
